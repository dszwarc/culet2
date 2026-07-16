from dataclasses import dataclass
from pathlib import Path
from typing import Any

from django.conf import settings
from openpyxl import load_workbook

from .utils import clean_text, is_populated


DEFAULT_REVIEW_WORKBOOK = (
    Path(settings.BASE_DIR)
    / "backups"
    / "ignore_old_culet_styles_and_parts.xlsx"
)


@dataclass(frozen=True)
class ReviewRow:
    old_id: int | None
    values: dict[str, Any]
    ignored: bool


def normalize_header(value: Any) -> str:
    """
    Normalize worksheet headers so values such as:

        Part Number
        PART NUMBER
        part_number

    can all be accessed consistently as 'part_number'.
    """
    header = clean_text(value).lower()
    return "_".join(header.split())


class ReviewWorkbook:
    """
    Reads the curated style and metal-part review workbook.

    Any populated value in a column named IGNORE causes that row to be skipped.
    """

    def __init__(self, path: str | Path | None = None):
        self.path = Path(path) if path else DEFAULT_REVIEW_WORKBOOK

        if not self.path.exists():
            raise FileNotFoundError(
                f"Review workbook was not found: {self.path}"
            )

    def read_sheet(self, sheet_name: str) -> list[ReviewRow]:
        workbook = load_workbook(
            filename=self.path,
            read_only=True,
            data_only=True,
        )

        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Worksheet '{sheet_name}' was not found in "
                    f"{self.path}. Available sheets: "
                    f"{', '.join(workbook.sheetnames)}"
                )

            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)

            try:
                raw_headers = next(rows)
            except StopIteration:
                return []

            headers = [normalize_header(value) for value in raw_headers]

            results: list[ReviewRow] = []

            for raw_row in rows:
                values = dict(zip(headers, raw_row, strict=False))

                if not any(value is not None for value in raw_row):
                    continue

                raw_id = values.get("id")

                try:
                    old_id = int(raw_id) if raw_id is not None else None
                except (TypeError, ValueError):
                    old_id = None

                results.append(
                    ReviewRow(
                        old_id=old_id,
                        values=values,
                        ignored=is_populated(values.get("ignore")),
                    )
                )

            return results

        finally:
            workbook.close()

    def approved_style_ids(self) -> set[int]:
        return {
            row.old_id
            for row in self.read_sheet("Styles")
            if not row.ignored and row.old_id is not None
        }

    def ignored_style_ids(self) -> set[int]:
        return {
            row.old_id
            for row in self.read_sheet("Styles")
            if row.ignored and row.old_id is not None
        }

    def approved_metal_part_ids(self) -> set[int]:
        return {
            row.old_id
            for row in self.read_sheet("Metal Parts")
            if not row.ignored and row.old_id is not None
        }

    def ignored_metal_part_ids(self) -> set[int]:
        return {
            row.old_id
            for row in self.read_sheet("Metal Parts")
            if row.ignored and row.old_id is not None
        }