from pathlib import Path

from django.conf import settings
from django.contrib.auth.models import User
from django.core.management.base import CommandError
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from culet.importer.base import BaseImportCommand
from culet.models import Department, Employee, LegacyRecordMap, Role


TEMP_PASSWORD = "temppassword"

REAL_CREDENTIAL_PATH = (
    Path(settings.BASE_DIR)
    / "backups"
    / "imported_employee_credentials.xlsx"
)

DRY_RUN_CREDENTIAL_PATH = (
    Path(settings.BASE_DIR)
    / "backups"
    / "imported_employee_credentials_dry_run.xlsx"
)


# Reviewed employee/login list from the July 19, 2026 Culet login sheet.
# Rows marked REMOVE are intentionally skipped.
# Rows marked Status are seeded by seed_job_statuses and are not employees.
# Rows marked TBD are imported as employees with no department.
EMPLOYEE_DEFINITIONS = [
    (1005, "Kleber", "Torres", "ktorres", "Jewelry", "Hourly"),
    (1008, "Darek", "Szumski", "dszumski", "Jewelry", "Hourly"),
    (1011, "Salvatore", "Aquino", "saquino", "Polishing", "Department Head"),
    (1015, "Jose", "Guevara", "jguevara", "Polishing", "Hourly"),
    (1018, "Ania", "Wasielewska", "awasielewska", "Quality Control", "Department Head"),
    (1024, "Andy", "Tester", "atester", "REMOVE", ""),
    (1025, "Greg", "Dusik", "gdusik", "Jewelry", "Manager"),
    (1029, "Agatha", "Szwarc", "aszwarc", "SUPER", "SUPER"),
    (1033, "Jaime", "Roa", "jroa", "Setting", "Hourly"),
    (1035, "Lenin", "Cordova", "lcordova", "Setting", "Hourly"),
    (1054, "Mach", "Ha", "mha", "Polishing", "Hourly"),
    (1095, "Marco", "Zumba", "mzumba", "Jewelry 37", "Hourly"),
    (1102, "Main", "Office 48", "moffice48", "REMOVE", ""),
    (1103, "Main", "Office 37", "moffice37", "REMOVE", ""),
    (1107, "Jennifer", "Mendez", "jmendez", "Setting", "Hourly"),
    (1110, "Robert", "Zandrowicz", "rzandrowicz", "Setting 37", "Hourly"),
    (1113, "Carlos", "Picon", "cpicon", "Jewelry", "Hourly"),
    (1115, "Jorge", "Garces", "jgarces", "Polishing", "Hourly"),
    (1117, "Cassandra", "Ferrone", "cferrone", "Office", "Manager"),
    (1119, "Peter", "Chen", "pchen", "Jewelry 37", "Hourly"),
    (1120, "Carlos", "Matute", "cmatute", "Jewelry", "Hourly"),
    (1124, "Diego", "Torres", "dtorres", "Setting", "Hourly"),
    (1132, "Luca", "Lombardi", "llombardi", "Jewelry", "Hourly"),
    (1136, "Hoa", "Mach", "hmach", "Polishing", "Hourly"),
    (1140, "Grigory", "Michnik", "gmichnik", "Contractor", ""),
    (1142, "Maria", "Barbosa", "mbarbosa", "Jewelry", "Hourly"),
    (1144, "Slawek", "Wieczorek", "swieczorek", "Setting 37", "Hourly"),
    (1151, "Nestor", "Vega", "nvega", "Jewelry 37", "Hourly"),
    (1155, "ALEX", "RBC", "arbc", "Contractor", ""),
    (1156, "RICARDO", "PASAN", "rpasan", "Polishing", "Hourly"),
    (1158, "Marek", "Kowalik", "mkowalik", "Contractor", ""),
    (1159, "WAITING ON", "MELEE", "wmelee", "Status", ""),
    (1160, "WAITING ON", "CENTER", "wcenter", "Status", ""),
    (1161, "WAITING ON", "FINDINGS", "wfindings", "Status", ""),
    (1162, "WAITING ON", "CASTINGS", "wcastings", "Status", ""),
    (1163, "Jake", "Kornecki", "jkornecki", "Jewelry", "Manager"),
    (1165, "Jose", "Andrade", "jandrade", "Polishing 37", "Hourly"),
    (1168, "WAITING ON", "OVERISSUE", "woverissue", "Status", ""),
    (1169, "CESAR", "FLORES", "cflores", "Jewelry", "Hourly"),
    (1171, "ARCHIL", "Molashkhia", "amolashkhia", "Polishing 37", "Department Head"),
    (1178, "Diego", "Granda", "dgranda", "Polishing", "Hourly"),
    (1183, "Edgar F", "Flores", "eflores", "Setting", "Hourly"),
    (1184, "Santiago", "Hungria", "shungria", "Polishing", "Hourly"),
    (1185, "ADAM", "MURPHY", "amurphy", "REMOVE", ""),
    (1187, "Cesar", "Garcia", "cgarcia", "REMOVE", ""),
    (1188, "Jose", "Garcia", "jgarcia", "REMOVE", ""),
    (1189, "Angel", "Mendez", "amendez", "REMOVE", ""),
    (1190, "Nohelia", "Punina", "npunina", "Polishing", "Hourly"),
    (1193, "Argentina", "Guevara", "aguevara", "Polishing", "Hourly"),
    (1195, "Oleg", "Selivanchikov", "oselivanchikov", "REMOVE", ""),
    (1206, "Luis", "Gonzalez", "lgonzalez", "Polishing", "Hourly"),
    (1207, "Edison", "Paredes", "eparedes", "Setting", "Hourly"),
    (1208, "Yackson", "Garcia", "ygarcia", "Polishing", "Hourly"),
    (1209, "Kris", "Kornecki", "kkornecki", "Jewelry 37", "Department Head"),
    (1210, "SAFE 3", "(Management Room)", "s3managementroom", "TBD", ""),
    (1211, "Safe 2", "(Office Near Entrance)", "s2officenearentrance", "TBD", ""),
    (1212, "Safe 1", "(Office Far Side)", "s1officefarside", "TBD", ""),
    (1214, "WAITING ON", "RECAST", "wrecast", "Status", ""),
    (1215, "Daniel", "Szwarc", "dszwarc", "SUPER", "super"),
    (1216, "Gregory", "Pestillo", "gpestillo", "Jewelry", "Hourly"),
    (1219, "Zurab", "Potskhverashvili", "zpotskhverashvili", "REMOVE", ""),
    (1221, "Larry", "Paredes", "lparedes", "Setting", "Hourly"),
    (1226, "Aly", "Niasse", "aniasse", "Jewelry", "Hourly"),
    (1227, "Paulina", "Mejia", "pmejia", "Quality Control", "Department Head"),
    (1228, "SAFE 4", "(POLISHING ROOM)", "s4polishingroom", "TBD", ""),
    (1230, "Shanna", "Matai", "smatai", "Office", "Manager"),
    (1231, "Diego", "Granda (LASER)", "dgrandalaser", "Jewelry", "Hourly"),
]


ROLE_NAME_MAP = {
    "hourly": "Hourly",
    "department head": "Department Head",
    "manager": "Manager",
    "super": "Super",
}


class Command(BaseImportCommand):
    help = (
        "Import the reviewed Culet employee/login list with departments, "
        "roles, and first-login password resets."
    )

    def add_arguments(self, parser):
        super().add_arguments(parser)
        parser.add_argument(
            "--reset-existing-passwords",
            action="store_true",
            help=(
                "Also reset mapped/existing users to 'temppassword'. "
                "Without this flag, only newly created users receive it."
            ),
        )

    def run_import(self, *args, **options):
        self.reset_existing_passwords = options["reset_existing_passwords"]
        self.credentials = []

        roles = {
            role.name.casefold(): role
            for role in Role.objects.filter(active=True)
        }
        departments = {
            department.name.casefold(): department
            for department in Department.objects.filter(active=True)
        }

        missing_roles = sorted(
            set(ROLE_NAME_MAP.values())
            - {role.name for role in roles.values()}
        )
        if missing_roles:
            raise CommandError(
                "Missing required roles: "
                + ", ".join(missing_roles)
                + ". Run `python manage.py seed_roles` first."
            )

        required_departments = {
            department_name
            for _, _, _, _, department_name, _ in EMPLOYEE_DEFINITIONS
            if department_name not in {"REMOVE", "Status", "TBD"}
        }
        missing_departments = sorted(
            name
            for name in required_departments
            if name.casefold() not in departments
        )
        if missing_departments:
            raise CommandError(
                "Missing required departments: "
                + ", ".join(missing_departments)
                + ". Run `python manage.py seed_departments` first."
            )

        total = len(EMPLOYEE_DEFINITIONS)
        for index, definition in enumerate(EMPLOYEE_DEFINITIONS, start=1):
            self.stats.processed += 1
            self.import_definition(
                definition=definition,
                roles=roles,
                departments=departments,
            )
            self.print_progress(index, total, "Employees")

        path = (
            DRY_RUN_CREDENTIAL_PATH
            if self.dry_run
            else REAL_CREDENTIAL_PATH
        )
        credentials = list(self.credentials)

        if self.dry_run:
            self.write_credentials_workbook(
                path=path,
                credentials=credentials,
                dry_run=True,
            )
        else:
            transaction.on_commit(
                lambda: self.write_credentials_workbook(
                    path=path,
                    credentials=credentials,
                    dry_run=False,
                ),
                using="default",
            )

    def import_definition(self, *, definition, roles, departments):
        (
            old_id,
            first_name,
            last_name,
            username,
            department_label,
            level_label,
        ) = definition

        if department_label == "REMOVE":
            self.stats.skipped += 1
            self.record_skipped(
                legacy_table="employee",
                legacy_id=old_id,
                message="Employee marked REMOVE in reviewed login sheet.",
            )
            self.row_message(f"SKIP employee {old_id}: marked REMOVE")
            return

        if department_label == "Status":
            self.stats.skipped += 1
            self.record_skipped(
                legacy_table="employee",
                legacy_id=old_id,
                message=(
                    "Row represents a JobStatus and is handled by "
                    "seed_job_statuses."
                ),
            )
            self.row_message(f"SKIP employee {old_id}: status row")
            return

        department = None
        if department_label != "TBD":
            department = departments[department_label.casefold()]

        normalized_level = level_label.strip().casefold()
        role_name = ROLE_NAME_MAP.get(normalized_level, "Hourly")
        role = roles[role_name.casefold()]

        with transaction.atomic(using="default"):
            employee = self.get_mapped_object(
                legacy_table="employee",
                legacy_id=old_id,
                model_class=Employee,
            )

            if employee is None:
                employee = Employee.objects.select_related("user").filter(
                    user__username__iexact=username
                ).first()

            created = employee is None
            if created:
                user = User.objects.create_user(
                    username=username,
                    password=TEMP_PASSWORD,
                    first_name=first_name,
                    last_name=last_name,
                    email="",
                    is_active=True,
                )
                employee = Employee.objects.create(
                    user=user,
                    department=department,
                    role=role,
                    clocked_in=False,
                    must_change_password=True,
                )
                action = LegacyRecordMap.ACTION_CREATED
                self.stats.created += 1
            else:
                user = employee.user
                changed = False

                user_values = {
                    "username": username,
                    "first_name": first_name,
                    "last_name": last_name,
                    "is_active": True,
                }
                user_changed_fields = []
                for field_name, desired_value in user_values.items():
                    if getattr(user, field_name) != desired_value:
                        setattr(user, field_name, desired_value)
                        user_changed_fields.append(field_name)

                if self.reset_existing_passwords:
                    user.set_password(TEMP_PASSWORD)
                    user_changed_fields.append("password")

                if user_changed_fields:
                    user.save(update_fields=list(dict.fromkeys(user_changed_fields)))
                    changed = True

                employee_changed_fields = []
                if employee.department_id != (department.id if department else None):
                    employee.department = department
                    employee_changed_fields.append("department")
                if employee.role_id != role.id:
                    employee.role = role
                    employee_changed_fields.append("role")
                if self.reset_existing_passwords and not employee.must_change_password:
                    employee.must_change_password = True
                    employee_changed_fields.append("must_change_password")

                if employee_changed_fields:
                    employee.save(update_fields=employee_changed_fields)
                    changed = True

                if changed:
                    action = LegacyRecordMap.ACTION_UPDATED
                    self.stats.updated += 1
                else:
                    action = LegacyRecordMap.ACTION_UNCHANGED
                    self.stats.unchanged += 1

            self.record_mapping(
                legacy_table="employee",
                legacy_id=old_id,
                target=employee,
                action=action,
            )

            self.credentials.append(
                {
                    "legacy_id": old_id,
                    "first_name": first_name,
                    "last_name": last_name,
                    "username": username,
                    "temporary_password": TEMP_PASSWORD,
                    "department": department.name if department else "",
                    "role": role.name,
                }
            )

            verb = "CREATE" if created else action.upper()
            self.row_message(
                f"{verb} employee {old_id}: {first_name} {last_name} "
                f"→ {username} / {role.name} / "
                f"{department.name if department else 'No department'}"
            )

    @staticmethod
    def write_credentials_workbook(*, path, credentials, dry_run):
        path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Employee Credentials"

        title = (
            "DRY-RUN PREVIEW — THESE PASSWORDS WILL NOT WORK"
            if dry_run
            else "Imported Employee Temporary Credentials"
        )
        worksheet.merge_cells("A1:G1")
        worksheet["A1"] = title
        worksheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        worksheet["A1"].fill = PatternFill(
            fill_type="solid",
            fgColor="C00000" if dry_run else "1F4E78",
        )

        worksheet.append([])
        worksheet.append(
            [
                "Legacy Employee ID",
                "First Name",
                "Last Name",
                "Username",
                "Temporary Password",
                "Department",
                "Role",
            ]
        )

        for credential in credentials:
            worksheet.append(
                [
                    credential["legacy_id"],
                    credential["first_name"],
                    credential["last_name"],
                    credential["username"],
                    credential["temporary_password"],
                    credential["department"],
                    credential["role"],
                ]
            )

        for cell in worksheet[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="5B9BD5")

        worksheet.freeze_panes = "A4"
        worksheet.auto_filter.ref = f"A3:G{worksheet.max_row}"

        widths = {
            "A": 20,
            "B": 22,
            "C": 26,
            "D": 24,
            "E": 22,
            "F": 22,
            "G": 20,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width

        workbook.save(path)