import csv
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import connections

from culet.models import Employee, LegacyRecordMap


class Command(BaseCommand):
    help = (
        "Export active employee names, Culet usernames, and "
        "legacy four-digit PINs to Excel or CSV."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            type=str,
            default="employee_credentials.xlsx",
            help=(
                "Output filename or path. Default: "
                "employee_credentials.xlsx"
            ),
        )

        parser.add_argument(
            "--format",
            choices=["xlsx", "csv"],
            default="xlsx",
            help="Export format. Default: xlsx",
        )

        parser.add_argument(
            "--hourly-only",
            action="store_true",
            help="Export only employees whose Culet role is Hourly.",
        )

        parser.add_argument(
            "--include-inactive-users",
            action="store_true",
            help=(
                "Include mapped employees whose Django user "
                "is inactive."
            ),
        )

    def handle(self, *args, **options):
        output_format = options["format"]
        hourly_only = options["hourly_only"]
        include_inactive_users = options[
            "include_inactive_users"
        ]

        output_path = self.get_output_path(
            requested_path=options["output"],
            output_format=output_format,
        )

        legacy_rows = self.get_legacy_employee_rows()

        export_rows = []
        stats = {
            "active_legacy": len(legacy_rows),
            "exported": 0,
            "invalid_pin": 0,
            "missing_mapping": 0,
            "missing_employee": 0,
            "inactive_user": 0,
            "non_hourly": 0,
        }

        for legacy_row in legacy_rows:
            export_row = self.build_export_row(
                legacy_row=legacy_row,
                stats=stats,
                hourly_only=hourly_only,
                include_inactive_users=include_inactive_users,
            )

            if export_row is not None:
                export_rows.append(export_row)

        export_rows.sort(
            key=lambda row: (
                row["last_name"].lower(),
                row["first_name"].lower(),
                row["username"].lower(),
            )
        )

        if not export_rows:
            raise CommandError(
                "No employee credentials qualified for export."
            )

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        if output_format == "xlsx":
            self.write_xlsx(
                output_path=output_path,
                rows=export_rows,
            )
        else:
            self.write_csv(
                output_path=output_path,
                rows=export_rows,
            )

        stats["exported"] = len(export_rows)

        self.print_summary(
            stats=stats,
            output_path=output_path,
        )

    def get_output_path(
        self,
        *,
        requested_path,
        output_format,
    ):
        output_path = Path(requested_path).expanduser()

        expected_suffix = f".{output_format}"

        if output_path.suffix.lower() != expected_suffix:
            output_path = output_path.with_suffix(
                expected_suffix
            )

        return output_path.resolve()

    def get_legacy_employee_rows(self):
        sql = """
            SELECT
                id,
                first_name,
                last_name,
                TRIM(pin) AS pin
            FROM employee
            WHERE is_active = 'Y'
            ORDER BY last_name, first_name, id
        """

        try:
            with connections["old_culet"].cursor() as cursor:
                cursor.execute(sql)

                columns = [
                    column[0]
                    for column in cursor.description
                ]

                return [
                    dict(zip(columns, values))
                    for values in cursor.fetchall()
                ]

        except Exception as exc:
            raise CommandError(
                "Could not read active employees from the "
                "old_culet database."
            ) from exc

    def build_export_row(
        self,
        *,
        legacy_row,
        stats,
        hourly_only,
        include_inactive_users,
    ):
        legacy_id = legacy_row["id"]
        first_name = (
            legacy_row["first_name"] or ""
        ).strip()
        last_name = (
            legacy_row["last_name"] or ""
        ).strip()
        pin = (
            legacy_row["pin"] or ""
        ).strip()

        # Require exactly four numeric characters.
        # This preserves leading zeros because the PIN remains text.
        if len(pin) != 4 or not pin.isdigit():
            stats["invalid_pin"] += 1

            self.stdout.write(
                self.style.WARNING(
                    f"SKIP INVALID PIN: legacy employee "
                    f"{legacy_id} — {first_name} {last_name}"
                )
            )
            return None

        mapping = (
            LegacyRecordMap.objects
            .filter(
                legacy_table="employee",
                legacy_id=legacy_id,
                object_id__isnull=False,
            )
            .order_by("-id")
            .first()
        )

        if mapping is None:
            stats["missing_mapping"] += 1

            self.stdout.write(
                self.style.WARNING(
                    f"SKIP NO MAPPING: legacy employee "
                    f"{legacy_id} — {first_name} {last_name}"
                )
            )
            return None

        employee = (
            Employee.objects
            .select_related(
                "user",
                "role",
            )
            .filter(pk=mapping.object_id)
            .first()
        )

        if employee is None:
            stats["missing_employee"] += 1

            self.stdout.write(
                self.style.WARNING(
                    f"SKIP MISSING EMPLOYEE: legacy employee "
                    f"{legacy_id} maps to Employee "
                    f"#{mapping.object_id}"
                )
            )
            return None

        if (
            not include_inactive_users
            and not employee.user.is_active
        ):
            stats["inactive_user"] += 1
            return None

        role_name = (
            employee.role.name
            if employee.role
            else ""
        )

        if hourly_only and role_name != "Hourly":
            stats["non_hourly"] += 1
            return None

        return {
            "first_name": (
                employee.user.first_name.strip()
                or first_name
            ),
            "last_name": (
                employee.user.last_name.strip()
                or last_name
            ),
            "username": employee.user.username,
            "pin": pin,
        }

    def write_xlsx(self, *, output_path, rows):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Alignment,
                Font,
                PatternFill,
            )
            from openpyxl.worksheet.table import (
                Table,
                TableStyleInfo,
            )
        except ImportError as exc:
            raise CommandError(
                "Excel export requires openpyxl. Install it with "
                "'pip install openpyxl', add it to requirements.txt, "
                "or run this command with '--format csv'."
            ) from exc

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Employee Credentials"

        headers = [
            "Employee Name",
            "Username",
            "PIN",
        ]

        worksheet.append(headers)

        for row in rows:
            employee_name = " ".join(
                part
                for part in [
                    row["first_name"],
                    row["last_name"],
                ]
                if part
            )

            worksheet.append(
                [
                    employee_name,
                    row["username"],
                    row["pin"],
                ]
            )

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )
        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # Force the PIN column to text so Excel keeps leading zeros.
        for cell in worksheet["C"][1:]:
            cell.number_format = "@"
            cell.alignment = Alignment(
                horizontal="center",
            )

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        worksheet.column_dimensions["A"].width = 30
        worksheet.column_dimensions["B"].width = 24
        worksheet.column_dimensions["C"].width = 12

        worksheet.row_dimensions[1].height = 22

        last_row = worksheet.max_row
        table_reference = f"A1:C{last_row}"

        table = Table(
            displayName="EmployeeCredentials",
            ref=table_reference,
        )

        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        worksheet.add_table(table)

        workbook.save(output_path)

    def write_csv(self, *, output_path, rows):
        # utf-8-sig makes the CSV easier to open directly in Excel.
        with output_path.open(
            mode="w",
            newline="",
            encoding="utf-8-sig",
        ) as csv_file:
            writer = csv.writer(csv_file)

            writer.writerow(
                [
                    "Employee Name",
                    "Username",
                    "PIN",
                ]
            )

            for row in rows:
                employee_name = " ".join(
                    part
                    for part in [
                        row["first_name"],
                        row["last_name"],
                    ]
                    if part
                )

                # ="0037" causes Excel to display and retain the
                # leading zeros when opening the CSV.
                excel_pin = f'="{row["pin"]}"'

                writer.writerow(
                    [
                        employee_name,
                        row["username"],
                        excel_pin,
                    ]
                )

    def print_summary(self, *, stats, output_path):
        self.stdout.write("")
        self.stdout.write(
            self.style.MIGRATE_HEADING(
                "Employee credential export summary"
            )
        )

        self.stdout.write(
            f"Active legacy employees: {stats['active_legacy']:,}"
        )
        self.stdout.write(
            f"Credentials exported:    {stats['exported']:,}"
        )
        self.stdout.write(
            f"Invalid or missing PIN:  {stats['invalid_pin']:,}"
        )
        self.stdout.write(
            f"Missing mapping:         {stats['missing_mapping']:,}"
        )
        self.stdout.write(
            f"Missing Employee:        {stats['missing_employee']:,}"
        )
        self.stdout.write(
            f"Inactive Culet users:    {stats['inactive_user']:,}"
        )
        self.stdout.write(
            f"Skipped non-hourly:      {stats['non_hourly']:,}"
        )

        self.stdout.write("")
        self.stdout.write(
            self.style.SUCCESS(
                f"Export created: {output_path}"
            )
        )

        self.stdout.write(
            self.style.WARNING(
                "This file contains employee passwords. Store it "
                "securely and delete it when it is no longer needed."
            )
        )