from datetime import datetime
from io import BytesIO
from urllib.parse import urlencode

from django.contrib.auth.models import User
from django.test import Client, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from .models import Employee, Role, TimeClock


class PayrollTestMixin:
    def setUp(self):
        self.viewer = User.objects.create_user(username="payroll-viewer", password="test")
        self.hourly_role = Role.objects.create(name="Hourly", requires_clock_in=True)
        self.excluded_role = Role.objects.create(name="Salary", requires_clock_in=False)
        self.john = self.make_employee("john", "John", "Smith", self.hourly_role)
        self.jane = self.make_employee("jane", "Jane", "Doe", self.hourly_role)
        self.excluded = self.make_employee("salary", "Sam", "Salary", self.excluded_role)
        self.client.force_login(self.viewer)

    @staticmethod
    def make_employee(username, first_name, last_name, role):
        user = User.objects.create_user(
            username=username, first_name=first_name, last_name=last_name
        )
        return Employee.objects.create(user=user, role=role)

    @staticmethod
    def aware(year, month, day, hour, minute=0):
        return timezone.make_aware(datetime(year, month, day, hour, minute))

    def make_entry(self, employee, start, end):
        return TimeClock.objects.create(employee=employee, clock_in=start, clock_out=end)


class TimeClockPayrollReturnTests(PayrollTestMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.entry = self.make_entry(
            self.john,
            self.aware(2026, 7, 27, 8),
            self.aware(2026, 7, 27, 16),
        )
        self.edit_url = reverse("culet:time_clock_edit", args=[self.entry.pk])
        self.payroll_url = reverse("culet:payroll_report")

    def valid_post(self, **extra):
        data = {
            "employee": self.john.pk,
            "clock_in": "2026-07-27T08:00",
            "clock_out": "2026-07-27T16:30",
        }
        data.update(extra)
        return data

    def test_safe_next_redirects_to_exact_filtered_payroll_url(self):
        query = urlencode(
            {"start_date": "2026-07-27", "end_date": "2026-08-09", "employee": self.john.pk}
        )
        next_url = f"{self.payroll_url}?{query}"
        response = self.client.post(
            f"{self.edit_url}?{urlencode({'next': next_url})}", self.valid_post(next=next_url)
        )
        self.assertRedirects(response, next_url, fetch_redirect_response=False)

    def test_missing_next_falls_back_to_payroll(self):
        response = self.client.post(self.edit_url, self.valid_post())
        self.assertRedirects(response, self.payroll_url, fetch_redirect_response=False)

    def test_external_next_is_rejected(self):
        response = self.client.post(
            f"{self.edit_url}?next=https%3A%2F%2Fevil.example%2Fsteal",
            self.valid_post(next="https://evil.example/steal"),
        )
        self.assertRedirects(response, self.payroll_url, fetch_redirect_response=False)

    def test_validation_errors_render_without_redirecting(self):
        response = self.client.post(
            self.edit_url,
            {"employee": "", "clock_in": "not-a-date", "clock_out": ""},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "This field is required")


class PayrollInlineTimeClockTests(PayrollTestMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.entry = self.make_entry(
            self.john,
            self.aware(2026, 8, 3, 8),
            self.aware(2026, 8, 3, 17),
        )
        self.query = urlencode(
            {
                "start_date": "2026-08-03",
                "end_date": "2026-08-09",
                "employee": self.john.pk,
            }
        )
        self.inline_url = (
            reverse("culet:payroll_timeclock_inline_edit", args=[self.entry.pk])
            + "?"
            + self.query
        )
        self.row_url = (
            reverse("culet:payroll_timeclock_row", args=[self.entry.pk])
            + "?"
            + self.query
        )
        self.htmx = {"HTTP_HX_REQUEST": "true"}

    def valid_post(self, clock_out="2026-08-03T16:00"):
        return {
            "employee": self.john.pk,
            "clock_in": "2026-08-03T08:00",
            "clock_out": clock_out,
        }

    def test_edit_returns_populated_compact_edit_row(self):
        response = self.client.get(self.inline_url, **self.htmx)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "reports/partials/payroll_timeclock_edit_row.html")
        self.assertContains(response, f'id="timeclock-row-{self.entry.pk}"')
        self.assertContains(response, 'value="2026-08-03T08:00"')
        self.assertContains(response, 'value="2026-08-03T17:00"')
        self.assertContains(response, "Save")
        self.assertContains(response, "Cancel")
        self.assertContains(response, 'hx-target="closest tr"', count=2)
        self.assertContains(response, 'hx-swap="outerHTML"', count=2)
        self.assertContains(response, 'hx-sync="this:drop"')
        self.assertContains(response, f'hx-select="#timeclock-row-{self.entry.pk}"')
        self.assertNotContains(response, "hx-disabled-elt")
        self.assertContains(response, self.query.replace("&", "&amp;"))

    def test_valid_post_updates_entry_and_returns_row_with_oob_totals(self):
        response = self.client.post(self.inline_url, self.valid_post(), **self.htmx)
        self.entry.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(timezone.localtime(self.entry.clock_out).hour, 16)
        self.assertContains(response, f'id="timeclock-row-{self.entry.pk}"')
        self.assertTemplateUsed(response, "reports/partials/payroll_inline_save_response.html")
        self.assertContains(response, "4:00 PM")
        self.assertContains(response, 'hx-target="closest tr"')
        self.assertContains(response, 'hx-swap="outerHTML"')
        self.assertContains(response, 'id="payroll-report-totals"')
        self.assertContains(response, 'hx-swap-oob="true"')

    def test_invalid_post_returns_edit_row_and_errors(self):
        response = self.client.post(
            self.inline_url,
            {"employee": self.john.pk, "clock_in": "invalid", "clock_out": ""},
            **self.htmx,
        )
        self.assertEqual(response.status_code, 422)
        self.assertTemplateUsed(response, "reports/partials/payroll_timeclock_edit_row.html")
        self.assertContains(response, "Enter a valid date/time", status_code=422)
        self.entry.refresh_from_db()
        self.assertEqual(timezone.localtime(self.entry.clock_out).hour, 17)

    def test_cancel_returns_unchanged_display_row(self):
        response = self.client.get(self.row_url, **self.htmx)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "reports/partials/payroll_timeclock_row.html")
        self.assertContains(response, "5:00 PM")
        self.assertNotContains(response, "Save")
        self.entry.refresh_from_db()
        self.assertEqual(timezone.localtime(self.entry.clock_out).hour, 17)

    def test_login_is_required_for_inline_editing(self):
        self.client.logout()
        self.assertEqual(self.client.get(self.inline_url, **self.htmx).status_code, 302)
        self.assertEqual(
            self.client.post(self.inline_url, self.valid_post(), **self.htmx).status_code,
            302,
        )

    def test_edit_from_41_to_40_updates_rounding_overtime_and_all_summaries(self):
        for day in range(4, 8):
            self.make_entry(
                self.john,
                self.aware(2026, 8, day, 8),
                self.aware(2026, 8, day, 16),
            )

        before = self.client.get(
            reverse("culet:payroll_report") + "?" + self.query
        )
        self.assertContains(before, "OT 1.00")

        response = self.client.post(self.inline_url, self.valid_post(), **self.htmx)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "OT 0.00")
        self.assertContains(response, "Rounded: 40.00")
        self.assertContains(response, "Overtime: 0.00")
        self.assertContains(response, "40.00")
        self.assertNotContains(response, "41.00")

    def test_non_htmx_request_preserves_standalone_editor(self):
        response = self.client.get(self.inline_url)
        self.assertRedirects(
            response,
            reverse("culet:time_clock_edit", args=[self.entry.pk]),
            fetch_redirect_response=False,
        )


class PayrollInlineTimeClockDeleteTests(PayrollTestMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.entry = self.make_entry(
            self.john,
            self.aware(2026, 8, 3, 8),
            self.aware(2026, 8, 3, 10),
        )
        self.query = urlencode(
            {
                "start_date": "2026-08-03",
                "end_date": "2026-08-09",
                "employee": self.john.pk,
            }
        )
        self.delete_url = (
            reverse("culet:payroll_timeclock_inline_delete", args=[self.entry.pk])
            + "?"
            + self.query
        )
        self.htmx = {"HTTP_HX_REQUEST": "true"}

    def test_payroll_row_has_confirmed_post_delete_action_and_filters(self):
        response = self.client.get(
            reverse("culet:payroll_report") + "?" + self.query
        )
        self.assertContains(response, "Edit")
        self.assertContains(response, "Delete")
        self.assertContains(response, 'hx-confirm="Delete this timeclock entry?"')
        self.assertContains(response, 'hx-target="closest tr"')
        self.assertContains(response, 'hx-swap="delete"')
        self.assertContains(response, self.query.replace("&", "&amp;"))

    def test_authorized_post_deletes_record_and_returns_oob_totals(self):
        response = self.client.post(self.delete_url, **self.htmx)
        self.assertEqual(response.status_code, 200)
        self.assertFalse(TimeClock.objects.filter(pk=self.entry.pk).exists())
        self.assertTemplateUsed(
            response,
            "reports/partials/payroll_inline_delete_response.html",
        )
        self.assertNotContains(response, f'id="timeclock-row-{self.entry.pk}"')
        self.assertContains(response, 'id="payroll-report-totals"')
        self.assertContains(response, 'hx-swap-oob="true"', count=5)
        self.assertContains(response, "0.00")

    def test_delete_requires_post(self):
        response = self.client.get(self.delete_url, **self.htmx)
        self.assertEqual(response.status_code, 405)
        self.assertTrue(TimeClock.objects.filter(pk=self.entry.pk).exists())

    def test_delete_requires_login(self):
        self.client.logout()
        response = self.client.post(self.delete_url, **self.htmx)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(TimeClock.objects.filter(pk=self.entry.pk).exists())

    def test_delete_is_csrf_protected(self):
        csrf_client = Client(enforce_csrf_checks=True)
        csrf_client.force_login(self.viewer)
        response = csrf_client.post(self.delete_url, **self.htmx)
        self.assertEqual(response.status_code, 403)
        self.assertTrue(TimeClock.objects.filter(pk=self.entry.pk).exists())

    def test_delete_recalculates_weekly_overtime_from_41_to_39(self):
        for day in range(4, 8):
            self.make_entry(
                self.john,
                self.aware(2026, 8, day, 8),
                self.aware(2026, 8, day, 16),
            )
        self.make_entry(
            self.john,
            self.aware(2026, 8, 3, 10),
            self.aware(2026, 8, 3, 17),
        )

        before = self.client.get(
            reverse("culet:payroll_report") + "?" + self.query
        )
        self.assertContains(before, "OT 1.00")

        response = self.client.post(self.delete_url, **self.htmx)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "OT 0.00")
        self.assertContains(response, "Rounded: 39.00")
        self.assertContains(response, "Overtime: 0.00")
        self.assertNotContains(response, "41.00")

    def test_open_timeclock_cannot_be_deleted(self):
        self.entry.clock_out = None
        self.entry.save(update_fields=["clock_out"])

        response = self.client.post(self.delete_url, **self.htmx)

        self.assertEqual(response.status_code, 409)
        self.assertContains(
            response,
            "Open TimeClock entries cannot be deleted",
            status_code=409,
        )
        self.assertTrue(TimeClock.objects.filter(pk=self.entry.pk).exists())

        payroll = self.client.get(
            reverse("culet:payroll_report") + "?" + self.query
        )
        self.assertContains(payroll, "Open TimeClock entries cannot be deleted")
        self.assertContains(payroll, "disabled")

    def test_already_deleted_entry_returns_404(self):
        self.entry.delete()
        response = self.client.post(self.delete_url, **self.htmx)
        self.assertEqual(response.status_code, 404)


class PayrollExcelTests(PayrollTestMixin, TestCase):
    def export(self, start="2026-07-29", end="2026-08-10", employee=None):
        params = {"start_date": start, "end_date": end}
        if employee:
            params["employee"] = employee.pk
        response = self.client.get(reverse("culet:payroll_excel"), params)
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        return response, workbook["Payroll"]

    def test_valid_xlsx_has_range_filename_repeating_weeks_and_numeric_totals(self):
        # Partial first/last weeks and 42.5 rounded hours in each of two weeks.
        for day in (29, 30, 31, 1, 2):
            month = 7 if day >= 29 else 8
            self.make_entry(
                self.john,
                self.aware(2026, month, day, 8),
                self.aware(2026, month, day, 16, 30),
            )
        for day in (3, 4, 5, 6, 7):
            self.make_entry(
                self.john,
                self.aware(2026, 8, day, 8),
                self.aware(2026, 8, day, 16, 30),
            )
        self.make_entry(
            self.jane, self.aware(2026, 7, 30, 8), self.aware(2026, 7, 30, 16)
        )
        self.make_entry(
            self.excluded, self.aware(2026, 7, 30, 8), self.aware(2026, 7, 30, 18)
        )

        response, sheet = self.export()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("payroll_2026-07-29_to_2026-08-10.xlsx", response["Content-Disposition"])
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            [
                "Employee", "Week 1 Time", "Week 1 Overtime",
                "Week 2 Time", "Week 2 Overtime", "Week 3 Time", "Week 3 Overtime",
                "Total Time", "Total Overtime",
            ],
        )
        rows = {row[0]: row[1:] for row in sheet.iter_rows(min_row=2, values_only=True)}
        self.assertEqual(rows["John Smith"], (42.5, 2.5, 42.5, 2.5, 0, 0, 85, 5))
        self.assertEqual(rows["Jane Doe"], (8, 0, 0, 0, 0, 0, 8, 0))
        self.assertEqual(list(rows["John Smith"][-2:]), [85, 5])
        self.assertEqual(
            rows["John Smith"][-2],
            sum(rows["John Smith"][0:-2:2]),
        )
        self.assertEqual(
            rows["John Smith"][-1],
            sum(rows["John Smith"][1:-2:2]),
        )
        self.assertEqual([cell.value for cell in sheet[1]][-2:], ["Total Time", "Total Overtime"])
        self.assertNotIn("Sam Salary", rows)
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet["B2"].number_format, "0.00")
        self.assertEqual(sheet.cell(row=2, column=sheet.max_column).number_format, "0.00")

    def test_totals_sum_weekly_values_without_offsetting_weekly_overtime(self):
        for day in range(27, 32):
            self.make_entry(
                self.john,
                self.aware(2026, 7, day, 8),
                self.aware(2026, 7, day, 17),
            )
        for day in range(3, 8):
            self.make_entry(
                self.john,
                self.aware(2026, 8, day, 8),
                self.aware(2026, 8, day, 15),
            )

        _response, sheet = self.export(start="2026-07-27", end="2026-08-09")

        self.assertEqual([cell.value for cell in sheet[1]][-2:], ["Total Time", "Total Overtime"])
        self.assertEqual(
            tuple(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0],
            ("John Smith", 45, 5, 35, 0, 80, 5),
        )

    def test_employee_filter_is_respected(self):
        self.make_entry(
            self.john, self.aware(2026, 8, 3, 8), self.aware(2026, 8, 3, 16)
        )
        self.make_entry(
            self.jane, self.aware(2026, 8, 3, 8), self.aware(2026, 8, 3, 16)
        )
        _response, sheet = self.export(start="2026-08-03", end="2026-08-09", employee=self.jane)
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet["A2"].value, "Jane Doe")

    def test_exactly_40_hours_has_no_overtime(self):
        for day in range(3, 8):
            self.make_entry(
                self.john, self.aware(2026, 8, day, 8), self.aware(2026, 8, day, 16)
            )
        _response, sheet = self.export(start="2026-08-03", end="2026-08-09")
        self.assertEqual(sheet["B2"].value, 40)
        self.assertEqual(sheet["C2"].value, 0)

    def test_export_uses_quarter_hour_payroll_rounding(self):
        self.make_entry(
            self.john,
            self.aware(2026, 8, 3, 8, 8),
            self.aware(2026, 8, 3, 16, 52),
        )
        _response, sheet = self.export(start="2026-08-03", end="2026-08-09")
        self.assertEqual(sheet["B2"].value, 8.5)

    def test_payroll_page_links_preserve_full_query_string(self):
        self.make_entry(
            self.john, self.aware(2026, 8, 3, 8), self.aware(2026, 8, 3, 16)
        )
        response = self.client.get(
            reverse("culet:payroll_report"),
            {"start_date": "2026-08-03", "end_date": "2026-08-09", "employee": self.john.pk},
        )
        self.assertContains(response, "Download Excel")
        self.assertContains(response, "next=")
